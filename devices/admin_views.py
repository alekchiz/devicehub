"""Кастомные страницы аналитики и загрузки внутри Django-admin."""
from datetime import timedelta

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.utils import timezone

from .analytics import device_uptime
from .models import Device, Location


@staff_member_required
def analytics_view(request):
    """Доступность киосков за период (неделя/месяц) из истории событий."""
    period = 'month' if request.GET.get('period') == 'month' else 'week'
    days = 30 if period == 'month' else 7
    sort = 'pct' if request.GET.get('sort') != 'hostname' else 'hostname'

    end = timezone.now()
    start = end - timedelta(days=days)
    total_seconds = (end - start).total_seconds()

    devices = Device.objects.filter(hostname__regex=r'^\d{3,}$').order_by('hostname')
    rows = []
    for d in devices:
        fraction = device_uptime(d, start, end)
        rows.append({
            'device': d,
            'pct': round(fraction * 100, 1),
            'online_h': round(total_seconds * fraction / 3600, 1),
            'offline_h': round(total_seconds * (1 - fraction) / 3600, 1),
        })

    if sort == 'pct':
        rows.sort(key=lambda r: (r['pct'], r['device'].hostname))
    else:
        rows.sort(key=lambda r: r['device'].hostname)

    avg = round(sum(r['pct'] for r in rows) / len(rows), 1) if rows else 0.0
    return render(request, 'admin/devices/analytics.html', {
        'title': 'Доступность киосков',
        'rows': rows,
        'period': period,
        'sort': sort,
        'period_days': days,
        'total': len(rows),
        'avg_uptime': avg,
        'healthy': sum(1 for r in rows if r['pct'] >= 99.0),
        'warning': sum(1 for r in rows if 90.0 <= r['pct'] < 99.0),
        'critical': sum(1 for r in rows if r['pct'] < 90.0),
    })


@staff_member_required
def import_view(request):
    """Загрузка Excel: SN клиента | расположение | количество осмотров."""
    from .views import _cell_text

    result = None
    header_tokens = {'sn', 'сн', 'номер', 'клиент', 'sn клиента', 'серийный номер'}

    if request.method == 'POST' and request.FILES.get('file'):
        from openpyxl import load_workbook

        stats = {'rows': 0, 'found': 0, 'updated': 0, 'not_found': [], 'no_exam': 0}
        workbook = load_workbook(request.FILES['file'], data_only=True, read_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            sn = _cell_text(row[0])
            location_name = row[1].strip() if len(row) > 1 and row[1] is not None else ''
            exam_cell = row[2] if len(row) > 2 else None
            if not sn or sn.lower() in header_tokens:
                continue

            stats['rows'] += 1
            exam_count = None
            if exam_cell is not None and exam_cell != '':
                try:
                    exam_count = int(float(exam_cell))
                except (TypeError, ValueError):
                    exam_count = None

            device = (Device.objects.filter(sn__iexact=sn).first()
                      or Device.objects.filter(hostname__iexact=sn).first())
            if not device:
                stats['not_found'].append(sn)
                continue

            stats['found'] += 1
            updates = {}
            if exam_count is not None:
                updates['exam_count'] = exam_count
            elif 'exam_count' not in updates:
                stats['no_exam'] += 1
            if location_name:
                location, _ = Location.objects.get_or_create(name=location_name)
                updates['location'] = location
            if updates:
                Device.objects.filter(pk=device.pk).update(**updates)
                stats['updated'] += 1

        result = stats

    return render(request, 'admin/devices/import.html', {
        'title': 'Загрузка данных киосков',
        'result': result,
    })
