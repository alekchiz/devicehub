"""Общие утилиты для формирования Excel-отчётов."""
import openpyxl
from django.http import HttpResponse

MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
HEADER_FILL = 'F1F5F9'


def new_workbook(sheet_title='Лист1'):
    """Возвращает (workbook, активный лист) с заданным заголовком листа."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws


def xlsx_response(wb, filename):
    """Обёртка над Workbook в виде HTTP-ответа для скачивания."""
    response = HttpResponse(content_type=MIMETYPE)
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


def style_header_row(ws, count):
    """Делает первую строку листа жирной с подложкой (шапка таблицы)."""
    from openpyxl.styles import Font, PatternFill
    font = Font(bold=True)
    fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    for cell in ws[1][:count]:
        cell.font = font
        cell.fill = fill


def autosize_columns(ws, min_width=10, max_width=42):
    """Подгоняет ширину колонок под содержимое."""
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        width = min(max(min_width, length + 2), max_width)
        ws.column_dimensions[col[0].column_letter].width = width