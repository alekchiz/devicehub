from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from devices.models import Device
from .models import Ticket


class TicketExportTests(TestCase):
    def setUp(self):
        self.device = Device.objects.create(hostname='123')
        self.tech1 = User.objects.create_user(username='t1', password='p')
        self.tech2 = User.objects.create_user(username='t2', password='p')
        self.admin = User.objects.create_user(username='admin', password='p')
        self.admin.profile.role = 'admin'
        self.admin.profile.save()

    def _ticket(self, created_by, assigned_to=None):
        return Ticket.objects.create(
            device=self.device,
            problem='Проблема',
            contact_name='Иван',
            contact_phone='+70000000000',
            created_by=created_by,
            assigned_to=assigned_to,
        )

    def _excel_ids(self, username):
        self.client.force_login(User.objects.get(username=username))
        resp = self.client.get(reverse('export_tickets'))
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        return [r[0] for r in wb.active.iter_rows(min_row=2, values_only=True) if r[0]]

    def test_technician_sees_own_created_and_assigned(self):
        own = self._ticket(self.tech1)
        foreign = self._ticket(self.tech2)
        assigned = self._ticket(self.tech1, assigned_to=self.tech2)

        ids = self._excel_ids('t1')
        self.assertIn(own.id, ids)
        self.assertIn(assigned.id, ids)   # автором предложенного, но назначен на tech2
        self.assertNotIn(foreign.id, ids)

    def test_technician_assigned_sees_ticket(self):
        own = self._ticket(self.tech1, assigned_to=self.tech2)
        ids = self._excel_ids('t2')
        self.assertIn(own.id, ids)

    def test_admin_sees_all(self):
        a = self._ticket(self.tech1)
        b = self._ticket(self.tech2)
        ids = self._excel_ids('admin')
        self.assertIn(a.id, ids)
        self.assertIn(b.id, ids)

    def test_export_requires_login(self):
        resp = self.client.get(reverse('export_tickets'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/accounts/login/', resp.url)


class TicketViewTests(TestCase):
    def setUp(self):
        self.device = Device.objects.create(hostname='123')
        self.tech1 = User.objects.create_user(username='t1', password='p')
        self.tech1.profile.role = 'technician'
        self.tech1.profile.save()
        self.tech2 = User.objects.create_user(username='t2', password='p')
        self.tech2.profile.role = 'technician'
        self.tech2.profile.save()
        self.admin = User.objects.create_user(username='admin', password='p')
        self.admin.profile.role = 'admin'
        self.admin.profile.save()

    def _ticket(self, created_by=None, assigned_to=None):
        return Ticket.objects.create(
            device=self.device,
            problem='Проблема',
            contact_name='Иван',
            contact_phone='+70000000000',
            created_by=created_by or self.tech1,
            assigned_to=assigned_to,
        )

    def test_list_requires_login(self):
        self.assertEqual(self.client.get(reverse('tickets_list')).status_code, 302)

    def test_create_requires_login(self):
        self.assertEqual(self.client.post(reverse('ticket_create'), {}).status_code, 302)

    def test_create_ticket_logs_activity(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('ticket_create'), {
            'device_id': self.device.id,
            'problem': 'Билл-лайт мигает',
            'contact_name': 'Иван',
            'contact_phone': '+70000000000',
        })
        ticket = Ticket.objects.get()
        self.assertEqual(ticket.created_by, self.admin)
        self.assertEqual(ticket.problem, 'Билл-лайт мигает')
        from .models import ActivityLog
        self.assertTrue(
            ActivityLog.objects.filter(action='create', object_id=ticket.id).exists()
        )

    def test_create_ignores_empty_problem(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('ticket_create'), {'device_id': self.device.id, 'problem': ''})
        self.assertEqual(Ticket.objects.count(), 0)

    def test_technician_cannot_edit_others_ticket(self):
        ticket = self._ticket(created_by=self.tech2)
        self.client.force_login(self.tech1)
        resp = self.client.post(reverse('ticket_edit', args=[ticket.pk]), {
            'problem': 'Меняет чужую заявку',
            'contact_name': 'X',
            'contact_phone': '+70000000000',
        })
        self.assertRedirects(resp, reverse('tickets_list'), fetch_redirect_response=False)
        ticket.refresh_from_db()
        self.assertNotEqual(ticket.problem, 'Меняет чужую заявку')

    def test_assign_sets_in_progress(self):
        ticket = self._ticket()
        self.client.force_login(self.admin)
        self.client.post(reverse('ticket_assign', args=[ticket.pk]), {'user_id': self.tech2.id})
        ticket.refresh_from_db()
        self.assertEqual(ticket.assigned_to, self.tech2)
        self.assertEqual(ticket.status, 'in_progress')

    def test_status_change_validates_choice(self):
        ticket = self._ticket()
        self.client.force_login(self.admin)
        self.client.post(reverse('ticket_change_status', args=[ticket.pk]), {'status': 'HACK'})
        ticket.refresh_from_db()
        self.assertEqual(ticket.status, 'created')
        self.client.post(reverse('ticket_change_status', args=[ticket.pk]), {'status': 'completed'})
        ticket.refresh_from_db()
        self.assertEqual(ticket.status, 'completed')

    def test_comment_created(self):
        ticket = self._ticket()
        self.client.force_login(self.admin)
        self.client.post(reverse('ticket_comment', args=[ticket.pk]), {'text': 'Позвонил клиенту'})
        self.assertTrue(ticket.comments.filter(text='Позвонил клиенту').exists())

    def test_technician_export_limited_to_own(self):
        own = self._ticket(created_by=self.tech1)
        foreign = self._ticket(created_by=self.tech2)
        self.client.force_login(self.tech1)
        resp = self.client.get(reverse('export_tickets'))
        wb = load_workbook(BytesIO(resp.content))
        ids = [r[0] for r in wb.active.iter_rows(min_row=2, values_only=True) if r[0]]
        self.assertIn(own.id, ids)
        self.assertNotIn(foreign.id, ids)
